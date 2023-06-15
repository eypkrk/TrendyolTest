from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from constants import globalTrendyolConstants
from time import sleep
from selenium.webdriver.common.action_chains import ActionChains
from datetime import date
from pathlib import Path
import pytest
import openpyxl

class TestTrendyol:
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalTrendyolConstants.URL)
        self.folderpath = str(date.today())
        Path(self.folderpath).mkdir(exist_ok=True)
            

    def teardown_method(self):
        self.driver.quit()

    def getUsers():
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet = excelFile["Sheet1"]
        totalRows = selectedSheet.max_row
        data = []
        for i in range(2,totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password) 
            data.append(tupleData)
        return data

    def waitMethod(self,locater):
        WebDriverWait(self.driver,3).until(ec.visibility_of_element_located((locater)))

    def btnOfWoman(self):
        #butonun üstüne gelme işlemi
        btnOn= self.driver.find_element(By.XPATH,"//div[@id='navigation-wrapper']/nav/ul/li[1]/a")
        actions = ActionChains(self.driver)
        actions.move_to_element(btnOn).perform()
        self.waitMethod((By.XPATH,"//div[@id='sub-nav-1']/div/div/div/div/ul/li/a"))

    def btnOfMan(self):
        btnOn = self.driver.find_element(By.XPATH,"//*[@id='navigation-wrapper']/nav/ul/li[2]/a")     
        actions = ActionChains(self.driver)
        actions.move_to_element(btnOn).perform()
        self.waitMethod((By.XPATH,"//*[@id='navigation-wrapper']/nav/ul/li[2]/a"))
    
    def crossAdvert(self):   
        btnCross = self.driver.find_element(By.ID,"Combined-Shape")
        btnCross.click()

    def test_buttonsOfWomans(self):
        self.crossAdvert()
        for i in range(13):
            i += 1
            self.btnOfWoman()
            btnLink = self.driver.find_element(By.XPATH,f"//*[@id='sub-nav-1']/div/div/div[1]/div/ul/li[{i}]/a")
            btnLink.click()
            self.driver.save_screenshot(f"{self.folderpath}/{i}woman.png")
        self.waitMethod((By.XPATH,"//*[@id='search-app']/div/div[1]/div[2]/div[1]/div[2]/div/div/div"))
        btnSort = self.driver.find_element(By.XPATH,"//div[@id='search-app']/div/div/div[2]/div/div[2]/div/div/div")
        btnSort.click()
        btnMin = self.driver.find_element(By.XPATH,"//div[@id='search-app']/div/div/div[2]/div/div[2]/div/ul/li[2]")
        btnMin.click()
        self.driver.save_screenshot(f"{self.folderpath}/{i}womanorder.png")

    def test_buttonsOfMan(self):
        self.crossAdvert()
        for j in range(1,14):
            self.btnOfMan()
            self.waitMethod((By.XPATH,f"//*[@id='sub-nav-2']/div/div/div[1]/div/ul/li[{j}]"))
            btnLink1 = self.driver.find_element(By.XPATH,f"//*[@id='sub-nav-2']/div/div/div[1]/div/ul/li[{j}]/a")
            btnLink1.click()
            self.driver.save_screenshot(f"{self.folderpath}/{j}man.png")
    
    def test_filter(self):
        self.crossAdvert()
        btnOn = self.driver.find_element(By.XPATH,"//*[@id='navigation-wrapper']/nav/ul/li[3]/a")
        actions = ActionChains(self.driver)
        actions.move_to_element(btnOn).perform()
        self.waitMethod((By.XPATH,"//*[@id='sub-nav-3']/div/div/div[1]/div/ul/li[1]/a"))
        btnLink = self.driver.find_element(By.XPATH,"//*[@id='sub-nav-3']/div/div/div[1]/div/ul/li[1]/a")
        btnLink.click()
        txtArea = self.driver.find_element(By.XPATH,"//*[@id='sticky-aggregations']/div/div[2]/div[2]/input")
        txtArea.send_keys("Civil Baby")
        btnFilter = self.driver.find_element(By.XPATH,"//*[@id='sticky-aggregations']/div/div[2]/div[3]/div/div/div/div/a")
        btnFilter.click()
        self.driver.save_screenshot(f"{self.folderpath}/1.png")
        sleep(2)
        btnSize = self.driver.find_element(By.XPATH,"//*[@id='sticky-aggregations']/div/div[3]/div[3]/div/div/div[5]/div/a")
        btnSize.click()
        actions.move_to_element(btnSize).perform()
        sleep(2)
        self.driver.save_screenshot(f"{self.folderpath}/2.png")
        btnImg = self.driver.find_element(By.XPATH,"//*[@id='search-app']/div/div[1]/div[2]/div[4]/div[1]/div/div[2]/div[1]/a")
        actions.move_to_element(btnImg).perform()
        self.waitMethod((By.XPATH,"//*[@id='search-app']/div/div[1]/div[2]/div[4]/div[1]/div/div[2]/div[1]/a"))
        btnImg.click()
        self.driver.save_screenshot(f"{self.folderpath}/3.png")
        btnSort = self.driver.find_element(By.XPATH,"//*[@id='search-app']/div/div[1]/div[2]/div[1]/div[2]/div")
        btnSort.click()
        self.driver.save_screenshot(f"{self.folderpath}/4.png")
        btnMax = self.driver.find_element(By.XPATH,"//*[@id='search-app']/div/div[1]/div[2]/div[1]/div[2]/div/ul/li[3]")
        btnMax.click()
        self.driver.save_screenshot(f"{self.folderpath}/5.png")

    def test_moreSeller(self):
        self.crossAdvert()
        actions = ActionChains(self.driver)
        searchArea = self.driver.find_element(By.XPATH,"//*[@id='sfx-discovery-search-suggestions']/div/div/input")
        actions.move_to_element(searchArea).perform()
        self.waitMethod((By.XPATH,"//*[@id='navigation-wrapper']/nav/ul/li[10]"))
        btnSeller = self.driver.find_element(By.XPATH,"//*[@id='navigation-wrapper']/nav/ul/li[10]")
        btnSeller.click()
        self.waitMethod((By.XPATH,"//*[@id='category-top-ranking']/div/div[1]/div/div[1]/div/div[1]"))
        btnAllCategory = self.driver.find_element(By.XPATH,"//*[@id='category-top-ranking']/div/div[1]/div/div[1]/div/div[1]")
        btnAllCategory.click()
        sleep(2)
        self.waitMethod((By.XPATH,"//div[@id='category-top-ranking']/div/div/div/div[2]/div/div/div[2]"))
        btnMan = self.driver.find_element(By.XPATH,"//div[@id='category-top-ranking']/div/div/div/div[2]/div/div/div[2]")
        btnMan.click()
        btnTshirt = self.driver.find_element(By.XPATH,"//*[@id='category-top-ranking']/div/div[3]/div/div[14]/a/div[1]")
        actions.move_to_element(btnTshirt).perform()
        self.driver.save_screenshot(f"{self.folderpath}/6.png")
        btnMore = self.driver.find_element(By.XPATH,"//*[@id='category-top-ranking']/div/div[2]/div/div[2]/button[2]")
        btnMore.click()
        self.driver.save_screenshot(f"{self.folderpath}/7.png")
        btnEvaluate = self.driver.find_element(By.XPATH,"//*[@id='category-top-ranking']/div/div[2]/div/div[2]/button[3]")
        btnEvaluate.click()
        self.driver.save_screenshot(f"{self.folderpath}/8.png")
    
    def test_flashProducts(self):
        self.crossAdvert()
        actions = ActionChains(self.driver)
        btnProducts = self.driver.find_element(By.XPATH,"//div[@id='navigation-wrapper']/nav/ul/li[11]/a")
        actions.move_to_element(btnProducts)
        btnProducts.click()
        self.driver.execute_script("window.scrollTo(0,775)")
        btnProduct = self.driver.find_element(By.XPATH,"//*[@id='search-app']/div/div[2]/div[2]/div[3]/div[1]/div/div[6]")
        actions.move_to_element(btnProduct)
        self.driver.save_screenshot(f"{self.folderpath}/9.png")
        sleep(10)
        self.waitMethod((By.XPATH,"//*[@id='search-app']/div/div[2]/div[2]/div[3]/div[1]/div/div[6]/div[1]/div[2]/button"))
        btnAdd = self.driver.find_element(By.XPATH,"//*[@id='search-app']/div/div[2]/div[2]/div[3]/div[1]/div/div[6]/div[1]/div[2]/button")
        btnAdd.click()
        self.driver.save_screenshot(f"{self.folderpath}/10.png")
        self.driver.execute_script("window.scrollTo(0,0)")
        self.waitMethod((By.XPATH,"//*[@id='account-navigation-container']/div/div[2]"))
        shoppingCart = self.driver.find_element(By.XPATH,"//*[@id='account-navigation-container']/div/div[2]")
        shoppingCart.click()
        self.driver.save_screenshot(f"{self.folderpath}/11.png")

    @pytest.mark.parametrize("username,password",getUsers())
    def test_account(self,username,password):
        self.crossAdvert()
        actions = ActionChains(self.driver)
        btnSignOn = self.driver.find_element(By.XPATH,"//*[@id='account-navigation-container']/div/div[1]/div[1]")
        actions.move_to_element(btnSignOn).perform()
        self.waitMethod((By.XPATH,"//*[@id='account-navigation-container']/div/div[1]/div[2]/div/div[1]"))
        btnSignIn = self.driver.find_element(By.XPATH,"//*[@id='account-navigation-container']/div/div[1]/div[2]/div/div[1]")
        btnSignIn.click()
        self.driver.save_screenshot(f"{self.folderpath}/12.png")
        sleep(5)
        self.waitMethod((By.XPATH,"//*[@id='login-email']"))
        userName = self.driver.find_element(By.ID,"login-email")
        userName.send_keys(username)
        passworD = self.driver.find_element(By.XPATH,"//*[@id='login-password-input']")
        passworD.send_keys(password)
        self.waitMethod((By.XPATH,"//*[@id='login-register']/div[3]/div[1]/form/div[2]/div/i"))
        btnEye = self.driver.find_element(By.XPATH,"//*[@id='login-register']/div[3]/div[1]/form/div[2]/div/i")
        btnEye.click()
        self.driver.save_screenshot(f"{self.folderpath}/{username}-13.png")
        btnLogin = self.driver.find_element(By.XPATH,"//*[@id='login-register']/div[3]/div[1]/form/button")
        btnLogin.click()
        errorMsg = self.driver.find_element(By.ID,"error-box-wrapper")
        assert errorMsg.text == "Lütfen geçerli bir e-posta adresi giriniz."
       
        
        
        
        
        
        